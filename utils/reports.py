"""
TaleemCheck AI - Report Generation Module
Generates downloadable Excel and PDF reports.
"""

import io
import pandas as pd
from datetime import datetime


def generate_excel_report(all_results: dict, student_summary: pd.DataFrame, df_flat: pd.DataFrame) -> bytes:
    """
    Generate a comprehensive Excel report with multiple sheets.
    Returns bytes for download.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )

        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Sheet 1: Class Summary
            summary_export = student_summary.copy()
            summary_export.index.name = "Rank"
            summary_export.to_excel(writer, sheet_name="Class Summary", index=True)

            # Sheet 2: Detailed Results
            df_flat.to_excel(writer, sheet_name="Detailed Results", index=False)

            # Sheet 3: Per-student sheets
            for student_name, results in all_results.items():
                rows = []
                for r in results:
                    rows.append({
                        "Question": r.get("question_number", ""),
                        "Question Text": r.get("question", "")[:100],
                        "Max Marks": r.get("max_marks", 0),
                        "Marks Awarded": r.get("marks_awarded", 0),
                        "Justification": r.get("marks_justification", ""),
                        "Strengths": "; ".join(r.get("strengths", [])),
                        "Missing Concepts": "; ".join(r.get("missing_concepts", [])),
                        "Feedback (EN)": r.get("feedback_english", ""),
                        "Feedback (Roman Urdu)": r.get("feedback_roman_urdu", ""),
                        "Improvement": "; ".join(r.get("improvement_suggestions", [])),
                    })
                df_student = pd.DataFrame(rows)
                sheet_name = student_name[:31]  # Excel sheet name limit
                df_student.to_excel(writer, sheet_name=sheet_name, index=False)

            # Style the Class Summary sheet
            ws = writer.sheets["Class Summary"]
            header_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

        output.seek(0)
        return output.read()

    except Exception as e:
        # Fallback: return simple CSV as bytes
        return df_flat.to_csv(index=False).encode("utf-8")


def generate_pdf_report(student_name: str, results: list, total_obtained: int, total_max: int) -> bytes:
    """
    Generate a PDF report for a single student using fpdf2.
    Returns bytes for download.
    """
    try:
        from fpdf import FPDF

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # ── Header ──────────────────────────────────────────────
        pdf.set_fill_color(26, 82, 118)  # Dark blue
        pdf.rect(0, 0, 210, 35, "F")
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 20)
        pdf.set_xy(10, 8)
        pdf.cell(0, 10, "TaleemCheck AI", ln=True, align="C")
        pdf.set_font("Helvetica", "", 11)
        pdf.set_xy(10, 20)
        pdf.cell(0, 8, "Student Evaluation Report", ln=True, align="C")

        pdf.set_text_color(0, 0, 0)
        pdf.ln(15)

        # ── Student Info ─────────────────────────────────────────
        percentage = round(total_obtained / max(total_max, 1) * 100, 1)
        grade = _assign_grade(percentage)

        pdf.set_font("Helvetica", "B", 14)
        pdf.set_fill_color(236, 240, 241)
        pdf.cell(0, 10, f"  Student: {student_name}", ln=True, fill=True)
        pdf.ln(2)
        pdf.set_font("Helvetica", "", 12)
        pdf.cell(90, 8, f"  Total Marks: {total_obtained} / {total_max}", border=1)
        pdf.cell(90, 8, f"  Percentage: {percentage}%  |  Grade: {grade}", border=1, ln=True)
        pdf.ln(5)

        report_date = datetime.now().strftime("%d %B %Y")
        pdf.set_font("Helvetica", "I", 10)
        pdf.set_text_color(128, 128, 128)
        pdf.cell(0, 6, f"  Report generated: {report_date}", ln=True)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(5)

        # ── Per Question Results ──────────────────────────────────
        for r in results:
            q_num = r.get("question_number", "Q")
            q_text = r.get("question", "")[:80]
            marks = r.get("marks_awarded", 0)
            max_m = r.get("max_marks", 0)

            # Question header
            pdf.set_fill_color(52, 152, 219)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 11)
            pdf.cell(0, 8, f"  {q_num}: {q_text}", ln=True, fill=True)

            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Helvetica", "B", 11)
            pdf.cell(0, 7, f"  Marks: {marks} / {max_m}", ln=True)

            pdf.set_font("Helvetica", "", 10)
            pdf.multi_cell(0, 6, f"  Justification: {r.get('marks_justification', '')}")

            # Strengths
            if r.get("strengths"):
                pdf.set_font("Helvetica", "B", 10)
                pdf.set_text_color(39, 174, 96)
                pdf.cell(0, 6, "  Strengths:", ln=True)
                pdf.set_font("Helvetica", "", 10)
                pdf.set_text_color(0, 0, 0)
                for s in r.get("strengths", []):
                    pdf.cell(0, 5, f"    • {s}", ln=True)

            # Missing Concepts
            if r.get("missing_concepts"):
                pdf.set_font("Helvetica", "B", 10)
                pdf.set_text_color(231, 76, 60)
                pdf.cell(0, 6, "  Missing Concepts:", ln=True)
                pdf.set_font("Helvetica", "", 10)
                pdf.set_text_color(0, 0, 0)
                for m in r.get("missing_concepts", []):
                    pdf.cell(0, 5, f"    • {m}", ln=True)

            # Feedback
            pdf.set_font("Helvetica", "I", 10)
            pdf.set_text_color(60, 60, 60)
            pdf.multi_cell(0, 6, f"  Feedback: {r.get('feedback_english', '')}")
            pdf.multi_cell(0, 6, f"  Roman Urdu: {r.get('feedback_roman_urdu', '')}")
            pdf.set_text_color(0, 0, 0)
            pdf.ln(4)

        # ── Footer ───────────────────────────────────────────────
        pdf.set_y(-20)
        pdf.set_font("Helvetica", "I", 9)
        pdf.set_text_color(128, 128, 128)
        pdf.cell(0, 6, "TaleemCheck AI  |  AI-Assisted Evaluation  |  Teacher-Approved Report", align="C")

        return pdf.output()

    except ImportError:
        return b"fpdf2 not installed. Please run: pip install fpdf2"
    except Exception as e:
        return f"PDF generation error: {str(e)}".encode()


def _assign_grade(percentage: float) -> str:
    if percentage >= 90: return "A+"
    elif percentage >= 80: return "A"
    elif percentage >= 70: return "B"
    elif percentage >= 60: return "C"
    elif percentage >= 50: return "D"
    else: return "F"
